import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import filedialog
import math 
import string


class Position:
    def __init__(self, name, count, hour_share, review_discipline):
        self.name = name
        self.count = count
        self.hour_share = hour_share
        self.discipline = review_discipline

class Review_Discipline:
    def __init__(self, name, hours):
        self.name = name
        self.total_hours = hours
    def __str__(self):
        s = "Name: " + str(self.name) + "\t"
        s += "Total Hours: " + str(self.total_hours)
        return s
        

class Permit():
    def __init__(self, per_year, name, sub_type_list2 = None):
        self.per_year = per_year
        self.type = name
        if sub_type_list2 == None:
            self.sub_type_list = []
        self.review_disciplines = []
    def __str__(self):
        s = "Type: " + str(self.type) + "\n"
        s += "Per Year: " + str(self.per_year)
        if len(self.review_disciplines) != 0:
            s += "\nReview Disciplines: \n"
            for disc in self.review_disciplines:
                s +=  str(disc) + "\n"

        if len(self.sub_type_list) != 0:
            s += "\nSubtypes:"
            for subtype in self.sub_type_list:
                s += "\n"+ str(subtype)
            s += "\n"
        return s



def main():
    test()

#https://stackoverflow.com/questions/7261936/convert-an-excel-or-spreadsheet-column-letter-to-its-number-in-pythonic-fashion
def col2num(col):
    x = 0
    y = 0
    for c in col:
        if c in string.ascii_letters:
            x = x * 26 + (ord(c.upper()) - ord('A')) + 1
        elif c.isdigit():
            y = y * 10 + int(c)
        else:
            print(col + "Invalid Cell Name")
            return (x,y)
    return (x,y)



    
def file_read_parser(name = "Template Input.xlsx", ask = True):
    filename = name
    if ask:
        filename=filedialog.askopenfilename() 
    #wb = Workbook()
    return load_workbook(filename, rich_text=True)

    sheet_names = thing.sheetnames
    typex = thing[sheet_names[1]]
    hoursx = thing[sheet_names[2]]
    staffing = thing[sheet_names[3]]





    categories = parse_types_x(typex)
    parse_hours_x(hoursx,categories)
    for i in categories["Building "][0].review_disciplines:
        print (i.name, i.total_hours)
    

def parse_types_x(typex, bottom= 62, category_cell = None, per_type_cell = "B2", per_year_cell = "C2"):
    """
    Docstring for parse_types_x
    
    :param typex: The pyxl worksheet variable we are parsing data from
    :param bottom: The bottommost row, to know when to stop reading
    :param category_cell: The cell that indicates where the complexity column begins, if it exists
    :param per_type_cell: The cell that indicates where the category type column begins
    :param per_year_cell: The cell that indicates where the per-year column begins
    """
    parse_type_generator = typex.iter_rows(min_row=row, max_row = bottom, max_col = column, values_only=True)
    categories = {}
    counter = 0
    for row in parse_type_generator:
        counter += 1
        Complexity = row[0] if complexity_exists else None
        Permit_type = row[0 + complexity_exists]
        Per_year = row[1 + complexity_exists]
        if Complexity == None and Per_year == None:
            categories[Permit_type]=[]
            most_recent = Permit_type

        elif Complexity == None:
            categories[most_recent].append(Permit(Per_year, Permit_type))
        
        elif Complexity != None:
            checker = 0
            for i in categories[most_recent]:
                # print(i.type)
                if i.type == Complexity:
                    i.per_year += Per_year
                    i.sub_type_list.append(Permit_type)
                    checker = 1
                    break
            if checker == 0:
                current = Permit(Per_year, Complexity)
                current.sub_type_list.append(Permit(Per_year, Permit_type))
                categories[most_recent].append(current)

    # print(counter)
    # print(categories.keys())
    #for i in categories.values():
        #print(len(i))
    return categories

def parse_hours_x(hoursx,categories, bottom = 50, row = 2, column = 6, min_col=1):
    """
    hoursx:
    """
    parse_hours_generator = hoursx.iter_rows(min_row=row, max_row = bottom, max_col=column, min_col=min_col, values_only=True)
    print("\n")
    for i in hoursx.iter_rows(max_row = 1, values_only=True, max_col =  6, min_col = 2):
        names = i
    print(names)
    #print(variable_names)
    counter = 0
    for row in parse_hours_generator:
        counter += 1
        if row[0] in categories:
            current = categories[row[0]]
        else:
            for i in range(len(current)):
                if current[i].type == row[0]:
                    for g in range(1,len(row)):
                        print(g)
                        current[i].review_disciplines.append(Review_Discipline(names[g-1],row[g]))
                    break

    return categories

# Permit('',),
building_list = [
    Permit(9.6,'Accessory Dwelling Unit/Adu'),
    Permit(5.4,'Accessory Structure/Garage'),
    Permit(0.6,'Apartments, Five Or More Units'),
    Permit(14.4,'Commercial'),
    Permit(2.4,'Commercial Kitchen Hood'),
    Permit(14.4,'Commercial Tenant Improvement'),
    Permit(4.8,'Cottage Unit'),
    Permit(23.4,'Deck'),
    Permit(12.0,'Deconstruction & Demolition'),
    Permit(62.4,'Fence'),
    Permit(1.8,'Manufactured Home'),
    Permit(2.4,'Manufactured Home Title Elimination'),
    Permit(4.2,'Over Or In Water Structure'),
    Permit(6.0,'Re-Roof Commercial'),
    Permit(75.6,'Re-Roof Residential'),
    Permit(21.0,'Residential Building Permit Application'),
    Permit(30.6,'Residential Interior Only Remodel'),
    Permit(39.6,'Residential Remodel'),
    Permit(9.0,'Residential Single Family  Building'),
    Permit(12.6,'Sign'),
    Permit(63.0,'Solar Permit'),
    Permit(1.8,'Tank Removal'),
    Permit(1.2,'Tent Temporary/Festival Use'),
    Permit(3.6,'Windows/Doors/Siding'),
    Permit(32.4,'Windows/Doors/Siding New Or Replaced'),
    Permit(1.2,'Wireless Facility/Cell Tower/Co-Location'),
    Permit(4.8,'Yard Irrigation Sprinklers')]
plumbing_mechanical_list =[
    Permit(21.0,'Mechanical And Gas Water Heaters, Commercial'),
    Permit(288.0,'Mechanical And Gas Water Heaters, Residential'),
    Permit(18.0,'Plumbing & Mechanical Permit'),
    Permit(24.6,'Plumbing And Electric Water Heaters')]
engineering_list = [
    Permit(0.6,'Site Development',),
    Permit(9.6,'Clear And Grade',),
    Permit(21.0,'Right-Of-Way Oversize Transport Application',),
    Permit(112.2,'Right-Of-Way Permit Application',),
    Permit(140.4,'Right-Of-Way Permit Franchise Application',),
    Permit(0.6,'Side Sewer Repair Or Replacement',)]
fire_list = [
    Permit(10.2,'Fire Sprinkler'),
    Permit(15.0,'Fire Alarm')
]
simple_permits = Permit(13.2,'Simple', [
    Permit(1.2,'Administrative Permit (Types 1 and 2)'),
    Permit(0.6,'Appeal Of Administrative Decision (heard by He)'),
    Permit(0.6,'Appeal Of Type 3 - Pc/He Decision (heard by Cc)'),
    Permit(4.2,'Boundary Line Adjustment (Type 1)'),
    Permit(0.6,'Conditional Use Permit (Type 4 -Cc)'),
    Permit(4.8,'Critical Area Exemption (Type 1)'),
    Permit(1.2,'Critical Area Permitted Alteration (Type 2)')
])
medium_permits = Permit(29.4,'Medium', [
    Permit(1.8,'Eligible Facilities Request (Type 1)'),
    Permit(4.2,'Floodplain Development'),
    Permit(0.6,'Legislative Amendment'),
    Permit(10.2,'Pre-Application Meeting Request'),
    Permit(10.8,'Shoreline Exemption Permit (Type 1)'),
    Permit(1.8,'Shoreline Substantial Development Permit')
])
hard_permits = Permit(4.2,'Hard', [
    Permit(1.8,'Site Plan Development'),
    Permit(2.4,'Unit Lot Subdivision')
])
hard_no_e_permits = Permit(0.6,'Hard - No Engineering', [
    Permit(0.6,'Wireless Service Facility (Type 1)')
])

planning_list = [simple_permits,medium_permits,hard_permits,hard_no_e_permits]
test_categories = {'Building': building_list, 'Plumbing / Mechanical': plumbing_mechanical_list, 'Engineering': engineering_list, 'Fire':fire_list}

def print_data(cate):
    print("There are " + str(len(cate)) +" catagories in the input")
    for category in cate:
        print(str(category) + " with " + str(len(cate[category]))+ " permit types:")
        for permit in cate[category]:
            print(permit)
        print("\n")

def test():
    wb = file_read_parser("Template Input.xlsx", False)
    print(wb)
    sheet_names = wb.sheetnames
    typex = wb[sheet_names[1]]
    hoursx = wb[sheet_names[2]]
    staffing = wb[sheet_names[3]]
    categories = parse_types_x(typex, complexity_exists= True)   
    categories = parse_hours_x(hoursx,categories)
    print_data(categories)

if __name__ == "__main__":
    main()