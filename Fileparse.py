import string
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import pandas as pd
import openpyxl
import math
from openpyxl import Workbook

# Hello peter. I am here to give you information.
# In order for this to work, you need to go to your VS code and upgrade the Python version to the latest version, 3.13.5.
# You can do this by going to View, opening the command pallete, searching "Python: Select Interpreter"
# Additionally, you will need to do "pip3 install openpyxl" in the terminal
# In order to get this shit running, you run UI.py

#types-x-per-interval
#hours-per-item-x-positions
#staffing-availibility

#xlsxwrite 

fte_per_person = 1872

class Position:
    def __init__(self, name, count, hour_share, review):
        self.name = name
        self.count = count
        self.hour_share = hour_share
        self.review_discipline = review

class Review_Discipline:
    def __init__(self, name, hours):
        self.name = name
        self.hours = hours
    def __equals__(a, b):
        return a.name == b.name

class Permit():
    def __init__(self, per_year, name):
        self.per_year = per_year
        self.type = name
        self.sub_type_list = []
    
    

def main():
    parser(False)

#https://stackoverflow.com/questions/7261936/convert-an-excel-or-spreadsheet-column-letter-to-its-number-in-pythonic-fashion
def col2num(col):
    x = 0
    y = 0
    for c in col:
        if c in string.ascii_letters:
            x = x * 26 + (ord(c.upper()) - ord('A')) + 1
        elif c.isdigit():
            y = y * 10 + int(c)
        else:
            print(col + "Invalid Cell Name")
            return (x,y)
    return (x,y)

def parser(ui = True, filename ="Template Input.xlsx"):
    print(col2num('A3'))
    print(col2num('AD15')[1])
    print(col2num("aw~/)0"))
    categories = {}
    
    if ui:
        filename=filedialog.askopenfilename()   
    from openpyxl import load_workbook
    wb = load_workbook(filename)
    for sheet in wb:
        print(sheet.title)
    wb.save('output.xlsx')

    # file = pd.ExcelFile(filename)
    # thing = pd.read_excel(file, sheet_name=None)
    # print(thing['staffing-availibility'])
    # print(type(thing))
    # output(thing)
    return 0

    # df = openpyxl.load_workbook(filename)
    # df.active
    # print(df.sheetnames)
    # sa = df["staffing-availibility"]
    # for row in sa.iter_rows():
    #     for cell in row:
    #         if isinstance(cell, openpyxl.cell.cell.MergedCell):
    #             print(cell)
    file = pd.ExcelFile(filename)
    thing = pd.read_excel(file, sheet_name=None, header=0)
    typesx = thing['types-x-per-interval']
    hoursx = thing['hours-per-item-x-positions']
    staffing = thing['staffing-availibility']
    counter = 0
    for index, row in typesx.iterrows():
        counter+=1
        if type(row['Catagory'])==float and math.isnan(row['Permits per year']):
            categories[row['Permit Type']]=[]
            most_recent = row['Permit Type']

        elif type(row['Catagory'])==float:
            categories[most_recent].append(Permit(row['Permits per year'], row['Permit Type']))

        elif type(row['Catagory'])!=float:
            checker = 0
            for i in categories[most_recent]:
                print(i.type)
                if i.type == row['Catagory']:
                    i.per_year += row['Permits per year']
                    i.sub_type_list.append(row['Permit Type'])
                    checker = 1
                    break
            if checker == 0:
                current = Permit(row['Permits per year'], row['Catagory'])
                current.sub_type_list.append(row['Permit Type'])
                categories[most_recent].append(current)
                
    print(counter)
    print(categories.keys())
    for i in categories.values():
        print(len(i))
    
    print(categories['Planning'][0].type)
    print(categories['Building '][0].type)
    print(categories['Planning'][1].sub_type_list)
    print(categories['Planning'][2].sub_type_list)
                


    
    
    
    
    #for i in thing.columns:
    #    print(i)
    #output(thing)
    #df = openpyxl.load_workbook(filename)
    #df.active
    #print(df.sheetnames)
    #sa = df["staffing-availibility"]
    #for row in sa.iter_rows():
    #    for cell in row:
    #        if isinstance(cell, openpyxl.cell.cell.MergedCell):
    #            print(cell)

def output(dataset):
    #tada!
    #https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.to_excel.html
        
     with pd.ExcelWriter('output.xlsx') as writer:
        dataset['types-x-per-interval'].to_excel(writer, sheet_name='types-x-per-interval')
        dataset['hours-per-item-x-positions'].to_excel(writer, sheet_name='hours-per-item-x-positions')
        dataset['staffing-availibility'].to_excel(writer, sheet_name='staffing-availibility')

if __name__ == "__main__":
    main()