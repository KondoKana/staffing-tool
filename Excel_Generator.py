from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import string

#https://stackoverflow.com/questions/7261936/convert-an-excel-or-spreadsheet-column-letter-to-its-number-in-pythonic-fashion
def col2num(col):
    x = 0
    y = 0
    for c in col:
        if c in string.ascii_letters:
            x = x * 26 + (ord(c.upper()) - ord('A')) + 1
        elif c.isdigit():
            y = y * 10 + int(c)
        else:
            print(col + "Invalid Cell Name")
            return (x,y)
    return (x,y)

def generate_input_file(fte_eq: int, 
                        permit_catagories: list[tuple[str, int]], 
                        review_disciplines: list[str], departments: 
                        list[tuple[str,str]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "overall-information"
    generate_overall_information(ws, fte_eq, departments)
    ws = wb.create_sheet(title="types-x-per-interval")
    generate_types_per_interval(ws, permit_catagories)
    ws = wb.create_sheet(title="hours-per-item")
    generate_hours_per(ws, permit_catagories, review_disciplines)

    wb.save('input-sheet.xlsx')

    return 0

def generate_overall_information(ws: openpyxl.worksheet, fte_eq : int, departments: list[tuple[str,str]]):
    ws["A1"] = "Departments"
    ws["B1"] = "Review Disciplines"
    ws["D1"] = "FTE"
    ws["E1"] = fte_eq

    for i in range (len(departments)):
        ws.cell(row = 2+i, column = 1, value =  departments[i][0])
        ws.cell(row = 2+i, column = 2, value =  departments[i][1])
    adjust_ws_width(ws)
    return ws



def generate_types_per_interval(ws: openpyxl.worksheet, permit_catagories: list[tuple[str, int]]):
    ws["A1"] = "Complexity"
    ws["B1"] = "Permit Type"
    ws["C1"] = "Permits per Year"

    height = len(permit_catagories)
    for catagory in permit_catagories:
        height += catagory[1]

    row = 2
    for catagory in permit_catagories:
        ws.cell(row = row, column = 2, value = catagory[0])
        row += 1
        for i in range(catagory[1]):
            ws.cell(row = row + i, column = 2, value = "-")
        row += catagory[1]
    adjust_ws_width(ws)
    return ws

def generate_hours_per(ws: openpyxl.worksheet, permit_catagories: list[tuple[str, int]], review_disciplines: list[str]):
    ws["A1"] = "Number of Review Disciplines"
    ws["B1"] = len(review_disciplines)
    ws["A2"] = "Permit Type"
    ws["B2"] = "Administrative Functions"

    row = 3
    for catagory in permit_catagories:
        ws.cell(row = row, column = 1, value = catagory[0])
        row += 1
        for i in range(catagory[1]):
            ws.cell(row = row + i, column = 1, value = "-")
        row += catagory[1]
    
    col = 3
    for discipline in review_disciplines:
        ws.cell(row = 2, column = col, value = discipline)
        col += 1
    adjust_ws_width(ws)
    return ws


def adjust_ws_width(ws: openpyxl.worksheet) :
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
    